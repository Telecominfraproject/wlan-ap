#!/usr/bin/env python3
"""
Step 5: Generate enhanced Excel from the annotated XML.

Reads ``gw3_current_datamodel.xml`` (produced by step 3) and generates an
Excel workbook with all TR-181 columns plus GW3 coverage annotations.

Sheets:
    - **Device-2**: Full data with 18 columns (TR-181 + GW3 annotations)
    - **Coverage Summary**: Per-subtree and per-service statistics
    - **Legend**: Color coding reference

Columns:
    A: Object Path  B: Item Path  C: Name  D: Type  E: Write  F: Description
    G: Object Default  H: Version
    I: Microservice  J: Mark  K: ODL Type  L: ODL Access  M: ODL Persistent
    N: ODL Volatile  O: ODL Default  P: ODL Source File  Q: Microservice Version
    R: Diff Details

Output:
    local/datamodel/gw3_current_datamodel.xlsx
"""

import argparse
import os
import re
import sys
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import LOCAL_DATAMODEL_DIR, GW3_NAMESPACE, ensure_output_dir
from xml_utils import (
    strip_ns, get_description_text, get_syntax_type, get_default_value,
    normalize_path,
)

GW3_NS = f'{{{GW3_NAMESPACE}}}'

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)

# Row type fills
OBJECT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
PARAM_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
COMMAND_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
EVENT_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
ARG_CONTAINER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
ARG_PARAM_FILL = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')

# Mark fills
MARK_FILLS = {
    'IMPLEMENTED': PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'),
    'NOT IMPLEMENTED': PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'),
    'DEFAULT ONLY': PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid'),
    'REPORT ONLY': PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid'),
    'NOT DEFINED': PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'),
    'MISMATCH': PatternFill(start_color='F4B084', end_color='F4B084', fill_type='solid'),
    'VENDOR EXTENSION': PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid'),
    'PRPL EXTENSION': PatternFill(start_color='D9D2E9', end_color='D9D2E9', fill_type='solid'),
}


def get_gw3_attr(elem, attr_name):
    """Get a gw3: attribute value from an element."""
    return elem.attrib.get(f'{GW3_NS}{attr_name}', '')


def process_enhanced_xml(xml_path):
    """Parse the enhanced XML and extract all rows with annotations."""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    model = None
    for child in root:
        if strip_ns(child.tag) == 'model':
            model = child
    if model is None:
        print("ERROR: No <model> element found.", file=sys.stderr)
        sys.exit(1)

    rows = []
    for obj in model:
        if strip_ns(obj.tag) != 'object':
            continue

        obj_name = obj.attrib.get('name', '')
        obj_access = obj.attrib.get('access', 'readOnly')
        obj_version = obj.attrib.get('version', '')
        obj_status = obj.attrib.get('status', '')
        obj_desc = get_description_text(obj.find('description'))
        obj_write = 'W' if obj_access == 'readWrite' else 'R'

        rows.append({
            'row_type': 'object',
            'obj_path': obj_name,
            'item_path': obj_name,
            'name': obj_name,
            'type': 'object',
            'write': obj_write,
            'description': obj_desc,
            'default': '-',
            'version': obj_version,
            'status': obj_status,
            'microservice': get_gw3_attr(obj, 'Microservice'),
            'mark': get_gw3_attr(obj, 'Mark'),
            'odl_type': get_gw3_attr(obj, 'ODL_Type'),
            'odl_access': get_gw3_attr(obj, 'ODL_Access'),
            'odl_persistent': get_gw3_attr(obj, 'ODL_Persistent'),
            'odl_volatile': get_gw3_attr(obj, 'ODL_Volatile'),
            'odl_default': get_gw3_attr(obj, 'ODL_Default'),
            'odl_source': get_gw3_attr(obj, 'ODL_Source_File'),
            'ms_version': get_gw3_attr(obj, 'Microservice_version'),
            'diff_details': get_gw3_attr(obj, 'DIFF_Details'),
        })

        for child in obj:
            ctag = strip_ns(child.tag)

            if ctag == 'parameter':
                p_name = child.attrib.get('name', '')
                p_access = child.attrib.get('access', 'readOnly')
                p_version = child.attrib.get('version', '')
                p_status = child.attrib.get('status', '')
                p_desc = get_description_text(child.find('description'))
                p_type = get_syntax_type(child)
                p_default = get_default_value(child)
                p_write = 'W' if p_access == 'readWrite' else 'R'

                rows.append({
                    'row_type': 'parameter',
                    'obj_path': obj_name,
                    'item_path': obj_name + p_name,
                    'name': p_name,
                    'type': p_type,
                    'write': p_write,
                    'description': p_desc,
                    'default': p_default,
                    'version': p_version,
                    'status': p_status,
                    'microservice': get_gw3_attr(child, 'Microservice'),
                    'mark': get_gw3_attr(child, 'Mark'),
                    'odl_type': get_gw3_attr(child, 'ODL_Type'),
                    'odl_access': get_gw3_attr(child, 'ODL_Access'),
                    'odl_persistent': get_gw3_attr(child, 'ODL_Persistent'),
                    'odl_volatile': get_gw3_attr(child, 'ODL_Volatile'),
                    'odl_default': get_gw3_attr(child, 'ODL_Default'),
                    'odl_source': get_gw3_attr(child, 'ODL_Source_File'),
                    'ms_version': get_gw3_attr(child, 'Microservice_version'),
                    'diff_details': get_gw3_attr(child, 'DIFF_Details'),
                })

            elif ctag == 'command':
                c_name = child.attrib.get('name', '')
                c_version = child.attrib.get('version', '')
                c_status = child.attrib.get('status', '')
                c_desc = get_description_text(child.find('description'))

                rows.append({
                    'row_type': 'command',
                    'obj_path': obj_name,
                    'item_path': obj_name + c_name,
                    'name': c_name,
                    'type': 'command',
                    'write': '-',
                    'description': c_desc,
                    'default': '-',
                    'version': c_version,
                    'status': c_status,
                    'microservice': get_gw3_attr(child, 'Microservice'),
                    'mark': get_gw3_attr(child, 'Mark'),
                    'odl_type': get_gw3_attr(child, 'ODL_Type'),
                    'odl_access': get_gw3_attr(child, 'ODL_Access'),
                    'odl_persistent': get_gw3_attr(child, 'ODL_Persistent'),
                    'odl_volatile': get_gw3_attr(child, 'ODL_Volatile'),
                    'odl_default': get_gw3_attr(child, 'ODL_Default'),
                    'odl_source': get_gw3_attr(child, 'ODL_Source_File'),
                    'ms_version': get_gw3_attr(child, 'Microservice_version'),
                    'diff_details': get_gw3_attr(child, 'DIFF_Details'),
                })

                # Add argument rows
                for section_name, section_tag in [('Input', 'input'), ('Output', 'output')]:
                    section = child.find(section_tag)
                    if section is not None:
                        rows.append({
                            'row_type': 'argument-container',
                            'obj_path': obj_name,
                            'item_path': '',
                            'name': f'\u21d2 {section_name}.',
                            'type': 'arguments',
                            'write': '-',
                            'description': f'{section_name} arguments.',
                            'default': '-', 'version': '', 'status': '',
                            'microservice': '', 'mark': '', 'odl_type': '',
                            'odl_access': '', 'odl_persistent': '',
                            'odl_volatile': '',
                            'odl_default': '', 'odl_source': '',
                            'ms_version': '', 'diff_details': '',
                        })
                        default_write = 'W' if section_name == 'Input' else 'R'
                        for arg in section:
                            arg_tag = strip_ns(arg.tag)
                            if arg_tag == 'parameter':
                                a_name = arg.attrib.get('name', '')
                                a_type = get_syntax_type(arg)
                                a_access = arg.attrib.get('access', '')
                                a_write = ('W' if a_access == 'readWrite' else 'R') if a_access else default_write
                                a_desc = get_description_text(arg.find('description'))
                                a_default = get_default_value(arg)
                                rows.append({
                                    'row_type': 'argument-parameter',
                                    'obj_path': obj_name,
                                    'item_path': '',
                                    'name': f'\u21d2 {a_name}',
                                    'type': a_type, 'write': a_write,
                                    'description': a_desc, 'default': a_default,
                                    'version': arg.attrib.get('version', ''),
                                    'status': arg.attrib.get('status', ''),
                                    'microservice': '', 'mark': '', 'odl_type': '',
                                    'odl_access': '', 'odl_persistent': '',
                                    'odl_volatile': '',
                                    'odl_default': '', 'odl_source': '',
                                    'ms_version': '', 'diff_details': '',
                                })

            elif ctag == 'event':
                e_name = child.attrib.get('name', '')
                e_version = child.attrib.get('version', '')
                e_status = child.attrib.get('status', '')
                e_desc = get_description_text(child.find('description'))

                rows.append({
                    'row_type': 'event',
                    'obj_path': obj_name,
                    'item_path': obj_name + e_name,
                    'name': e_name,
                    'type': 'event',
                    'write': '-',
                    'description': e_desc,
                    'default': '-',
                    'version': e_version,
                    'status': e_status,
                    'microservice': get_gw3_attr(child, 'Microservice'),
                    'mark': get_gw3_attr(child, 'Mark'),
                    'odl_type': get_gw3_attr(child, 'ODL_Type'),
                    'odl_access': get_gw3_attr(child, 'ODL_Access'),
                    'odl_persistent': get_gw3_attr(child, 'ODL_Persistent'),
                    'odl_volatile': get_gw3_attr(child, 'ODL_Volatile'),
                    'odl_default': get_gw3_attr(child, 'ODL_Default'),
                    'odl_source': get_gw3_attr(child, 'ODL_Source_File'),
                    'ms_version': get_gw3_attr(child, 'Microservice_version'),
                    'diff_details': get_gw3_attr(child, 'DIFF_Details'),
                })

                # Event argument rows
                has_args = False
                for ec in child:
                    ectag = strip_ns(ec.tag)
                    if ectag == 'parameter':
                        if not has_args:
                            rows.append({
                                'row_type': 'argument-container',
                                'obj_path': obj_name, 'item_path': '',
                                'name': '\u21d2 Output.', 'type': 'arguments',
                                'write': '-', 'description': 'Output arguments.',
                                'default': '-', 'version': '', 'status': '',
                                'microservice': '', 'mark': '', 'odl_type': '',
                                'odl_access': '', 'odl_persistent': '',
                                'odl_volatile': '',
                                'odl_default': '', 'odl_source': '',
                                'ms_version': '', 'diff_details': '',
                            })
                            has_args = True
                        a_name = ec.attrib.get('name', '')
                        a_type = get_syntax_type(ec)
                        a_desc = get_description_text(ec.find('description'))
                        a_default = get_default_value(ec)
                        rows.append({
                            'row_type': 'argument-parameter',
                            'obj_path': obj_name, 'item_path': '',
                            'name': f'\u21d2 {a_name}', 'type': a_type,
                            'write': 'R', 'description': a_desc,
                            'default': a_default,
                            'version': ec.attrib.get('version', ''),
                            'status': ec.attrib.get('status', ''),
                            'microservice': get_gw3_attr(ec, 'Microservice'),
                            'mark': get_gw3_attr(ec, 'Mark'),
                            'odl_type': get_gw3_attr(ec, 'ODL_Type'),
                            'odl_access': get_gw3_attr(ec, 'ODL_Access'),
                            'odl_persistent': get_gw3_attr(ec, 'ODL_Persistent'),
                            'odl_volatile': get_gw3_attr(ec, 'ODL_Volatile'),
                            'odl_default': get_gw3_attr(ec, 'ODL_Default'),
                            'odl_source': get_gw3_attr(ec, 'ODL_Source_File'),
                            'ms_version': get_gw3_attr(ec, 'Microservice_version'),
                            'diff_details': get_gw3_attr(ec, 'DIFF_Details'),
                        })

    return rows


def generate_excel(rows, output_path):
    """Generate the enhanced Excel workbook."""
    wb = Workbook()

    # --------------- Device-2 sheet ---------------
    ws = wb.active
    ws.title = "Device-2"

    headers = [
        'Object Path', 'Item Path', 'Name', 'Type', 'Write',
        'Description', 'Object Default', 'Version',
        'Microservice', 'Mark', 'ODL Type', 'ODL Access', 'ODL Persistent',
        'ODL Volatile', 'ODL Default', 'ODL Source File', 'Microservice Version',
        'Diff Details',
    ]
    col_widths = [40, 40, 20, 18, 7, 80, 16, 9, 22, 18, 14, 16, 13, 12, 16, 25, 18, 30]

    for ci, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{len(rows) + 1}'

    # Statistics tracking
    mark_counts = {}
    service_counts = {}
    subtree_stats = {}

    for row_idx, row_data in enumerate(rows, 2):
        row_type = row_data['row_type']
        status = row_data.get('status', '')
        mark = row_data.get('mark', '')

        # Determine row fill
        if row_type == 'object':
            fill = OBJECT_FILL
        elif row_type == 'command':
            fill = COMMAND_FILL
        elif row_type == 'event':
            fill = EVENT_FILL
        elif row_type == 'argument-container':
            fill = ARG_CONTAINER_FILL
        elif row_type in ('argument-parameter', 'argument-object'):
            fill = ARG_PARAM_FILL
        else:
            fill = PARAM_FILL

        values = [
            row_data['obj_path'], row_data['item_path'], row_data['name'],
            row_data['type'], row_data['write'], row_data['description'],
            row_data['default'], row_data['version'],
            row_data['microservice'], mark, row_data['odl_type'],
            row_data['odl_access'], row_data['odl_persistent'],
            row_data['odl_volatile'], row_data['odl_default'],
            row_data['odl_source'], row_data['ms_version'],
            row_data['diff_details'],
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=ci, value=val)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = Alignment(vertical='top', wrap_text=(ci in (1, 2, 3, 6, 18)))

        # Description cell special formatting
        desc_cell = ws.cell(row=row_idx, column=6)
        desc_text = row_data['description']
        if len(desc_text) > 32000:
            desc_cell.value = desc_text[:32000] + '... [truncated]'
        desc_cell.font = Font(name='Calibri', size=9)

        # Name cell formatting
        name_cell = ws.cell(row=row_idx, column=3)
        if row_type in ('object',):
            name_cell.font = Font(name='Calibri', bold=True, size=10)
        elif row_type == 'command':
            name_cell.font = Font(name='Calibri', bold=True, size=10, color='2E7D32')
        elif row_type == 'event':
            name_cell.font = Font(name='Calibri', bold=True, size=10, color='E65100')
        elif row_type == 'argument-container':
            name_cell.font = Font(name='Calibri', italic=True, size=10, color='666666')

        # Mark cell color coding
        mark_cell = ws.cell(row=row_idx, column=10)
        if mark in MARK_FILLS:
            mark_cell.fill = MARK_FILLS[mark]
        mark_cell.font = Font(name='Calibri', size=10, bold=True)

        # Track statistics for non-argument rows with marks
        if mark and row_type not in ('argument-container', 'argument-parameter', 'argument-object'):
            mark_counts[mark] = mark_counts.get(mark, 0) + 1

            svc = row_data.get('microservice', '')
            if svc:
                if svc not in service_counts:
                    service_counts[svc] = {'total': 0, 'IMPLEMENTED': 0, 'NOT IMPLEMENTED': 0,
                                           'NOT DEFINED': 0, 'MISMATCH': 0, 'VENDOR EXTENSION': 0, 'PRPL EXTENSION': 0}
                service_counts[svc]['total'] += 1
                service_counts[svc][mark] = service_counts[svc].get(mark, 0) + 1

            # Subtree stats
            item_path = row_data.get('item_path', '')
            if item_path.startswith('Device.'):
                parts = item_path.split('.')
                if len(parts) >= 2:
                    subtree = f'Device.{parts[1]}.'
                    if subtree not in subtree_stats:
                        subtree_stats[subtree] = {'total': 0, 'IMPLEMENTED': 0, 'NOT IMPLEMENTED': 0,
                                                  'NOT DEFINED': 0, 'MISMATCH': 0, 'VENDOR EXTENSION': 0, 'PRPL EXTENSION': 0}
                    subtree_stats[subtree]['total'] += 1
                    subtree_stats[subtree][mark] = subtree_stats[subtree].get(mark, 0) + 1

        if row_idx % 5000 == 0:
            print(f"  Written {row_idx - 1} rows...")

    # --------------- Coverage Summary sheet ---------------
    ws_sum = wb.create_sheet("Coverage Summary", 0)

    ws_sum.cell(row=1, column=1, value="GW3 TR-181 Data Model Coverage Report").font = Font(
        name='Calibri', bold=True, size=14)
    ws_sum.merge_cells('A1:F1')

    total = sum(mark_counts.values())
    implemented = mark_counts.get('IMPLEMENTED', 0) + mark_counts.get('MISMATCH', 0)
    not_impl = mark_counts.get('NOT IMPLEMENTED', 0)
    default_only = mark_counts.get('DEFAULT ONLY', 0)
    report_only = mark_counts.get('REPORT ONLY', 0)
    not_def = mark_counts.get('NOT DEFINED', 0)
    vendor = mark_counts.get('VENDOR EXTENSION', 0)
    prpl_ext = mark_counts.get('PRPL EXTENSION', 0)
    coverage_pct = (implemented + not_impl + default_only + report_only) / total * 100 if total else 0
    impl_pct = implemented / total * 100 if total else 0

    row = 3
    ws_sum.cell(row=row, column=1, value="Overall Status").font = Font(bold=True, size=12)
    row += 1
    overall_data = [
        ("Total TR-181 items", total),
        ("IMPLEMENTED (action callbacks)", implemented),
        ("NOT IMPLEMENTED (ODL defined, no actions)", not_impl),
        ("DEFAULT ONLY (readOnly, no callbacks)", default_only),
        ("REPORT ONLY (readWrite, C writes, no callbacks)", report_only),
        ("NOT DEFINED (not in ODL)", not_def),
        ("MISMATCH", mark_counts.get('MISMATCH', 0)),
        ("VENDOR EXTENSION (X_ prefix)", vendor),
        ("PRPL EXTENSION (non-X_ additions)", prpl_ext),
        ("Implementation rate", f"{impl_pct:.1f}%"),
        ("Coverage (defined in ODL)", f"{coverage_pct:.1f}%"),
    ]
    for label, value in overall_data:
        ws_sum.cell(row=row, column=1, value=label).font = Font(name='Calibri', size=10)
        cell = ws_sum.cell(row=row, column=2, value=value)
        cell.font = Font(name='Calibri', size=10, bold=True)
        cell.alignment = Alignment(horizontal='right')
        row += 1

    # Per-subtree table
    row += 1
    ws_sum.cell(row=row, column=1, value="Coverage by TR-181 Subtree").font = Font(bold=True, size=12)
    row += 1
    sub_headers = ["Subtree", "Total", "Implemented", "Not Implemented", "Not Defined", "Coverage %"]
    for ci, h in enumerate(sub_headers, 1):
        cell = ws_sum.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for subtree, st in sorted(subtree_stats.items(), key=lambda x: -x[1]['total']):
        t = st['total']
        imp = st.get('IMPLEMENTED', 0) + st.get('MISMATCH', 0)
        ni = st.get('NOT IMPLEMENTED', 0)
        nd = st.get('NOT DEFINED', 0)
        pct = (imp + ni) / t * 100 if t else 0

        ws_sum.cell(row=row, column=1, value=subtree).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=2, value=t).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=3, value=imp).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=4, value=ni).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=5, value=nd).font = Font(name='Calibri', size=10)
        pct_cell = ws_sum.cell(row=row, column=6, value=f"{pct:.1f}%")
        pct_cell.font = Font(name='Calibri', size=10, bold=True)
        if pct >= 80:
            pct_cell.fill = MARK_FILLS['IMPLEMENTED']
        elif pct >= 40:
            pct_cell.fill = MARK_FILLS['NOT IMPLEMENTED']
        else:
            pct_cell.fill = MARK_FILLS['NOT DEFINED']
        for ci in range(1, 7):
            ws_sum.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    # Per-service table
    row += 1
    ws_sum.cell(row=row, column=1, value="Coverage by Microservice").font = Font(bold=True, size=12)
    row += 1
    svc_headers = ["Service", "Total", "Implemented", "Not Implemented", "Not Defined"]
    for ci, h in enumerate(svc_headers, 1):
        cell = ws_sum.cell(row=row, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    row += 1

    for svc, sc in sorted(service_counts.items(), key=lambda x: -x[1]['total']):
        ws_sum.cell(row=row, column=1, value=svc).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=2, value=sc['total']).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=3, value=sc.get('IMPLEMENTED', 0) + sc.get('MISMATCH', 0)).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=4, value=sc.get('NOT IMPLEMENTED', 0)).font = Font(name='Calibri', size=10)
        ws_sum.cell(row=row, column=5, value=sc.get('NOT DEFINED', 0)).font = Font(name='Calibri', size=10)
        for ci in range(1, 6):
            ws_sum.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    # Column widths
    ws_sum.column_dimensions['A'].width = 50
    ws_sum.column_dimensions['B'].width = 18
    ws_sum.column_dimensions['C'].width = 15
    ws_sum.column_dimensions['D'].width = 18
    ws_sum.column_dimensions['E'].width = 15
    ws_sum.column_dimensions['F'].width = 14

    # --------------- Legend sheet ---------------
    ws_leg = wb.create_sheet("Legend")
    legend_data = [
        ['Mark', 'Color', 'Meaning'],
        ['IMPLEMENTED', 'Green', 'ODL entry exists with action callbacks (read/write handlers)'],
        ['NOT IMPLEMENTED', 'Yellow', 'ODL entry exists but has no action callbacks and no confirmed C writes'],
        ['DEFAULT ONLY', 'Gray', 'readOnly param in ODL without callbacks; AMX serves default/init value only'],
        ['REPORT ONLY', 'Steel', 'readWrite param with confirmed C writes but no action callbacks; value available for reading'],
        ['NOT DEFINED', 'Red', 'No ODL definition found for this TR-181 item'],
        ['MISMATCH', 'Orange', 'ODL entry exists but type or access differs from TR-181 spec'],
        ['VENDOR EXTENSION', 'Blue', 'ODL entry with X_ prefix, not in TR-181 standard (vendor-specific)'],
        ['PRPL EXTENSION', 'Purple', 'ODL entry without X_ prefix, not in TR-181 standard (prplOS addition)'],
    ]
    legend_fills = [HEADER_FILL] + [MARK_FILLS.get(row[0], None) for row in legend_data[1:]]
    for ri, row_vals in enumerate(legend_data, 1):
        for ci, val in enumerate(row_vals, 1):
            cell = ws_leg.cell(row=ri, column=ci, value=val)
            if ri == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            else:
                cell.font = Font(name='Calibri', size=10)
                if ci == 2 and legend_fills[ri - 1]:
                    cell.fill = legend_fills[ri - 1]
            cell.border = THIN_BORDER
    ws_leg.column_dimensions['A'].width = 22
    ws_leg.column_dimensions['B'].width = 12
    ws_leg.column_dimensions['C'].width = 80

    wb.save(output_path)
    print(f"\nExcel saved: {output_path}")
    print(f"  Total rows: {len(rows)}")
    print(f"  Mark distribution: {mark_counts}")


def main():
    parser = argparse.ArgumentParser(description='Generate enhanced Excel from annotated XML')
    parser.add_argument('--input', '-i', help='Input enhanced XML (default: gw3_current_datamodel.xml)')
    parser.add_argument('--output-dir', '-o', default=LOCAL_DATAMODEL_DIR)
    args = parser.parse_args()

    ensure_output_dir(args.output_dir)

    xml_path = args.input or os.path.join(LOCAL_DATAMODEL_DIR, 'gw3_current_datamodel.xml')
    if not os.path.exists(xml_path):
        print(f"ERROR: Enhanced XML not found: {xml_path}", file=sys.stderr)
        print("Run step3 first.", file=sys.stderr)
        sys.exit(1)

    output_path = os.path.join(args.output_dir, 'gw3_current_datamodel.xlsx')
    print(f"Input:  {xml_path}")
    print(f"Output: {output_path}")

    rows = process_enhanced_xml(xml_path)
    generate_excel(rows, output_path)


if __name__ == '__main__':
    main()
