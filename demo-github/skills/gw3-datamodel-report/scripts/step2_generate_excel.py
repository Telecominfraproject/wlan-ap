#!/usr/bin/env python3
"""
Step 2: Generate a standard Excel spreadsheet from TR-181 XML.

Parses the BBF TR-181 data-model XML and produces an Excel workbook with:
- **Device-2** sheet: all objects, parameters, commands, events, arguments
- **Data Types** sheet: all ``<dataType>`` definitions
- **References** sheet: all ``<bibliography>`` references
- **Legend** sheet: row type color legend

Columns:
    A: Object Path  B: Item Path  C: Name  D: Type  E: Write
    F: Description  G: Object Default  H: Version

Output:
    local/datamodel/tr181_definition.xlsx
"""

import argparse
import os
import sys
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from config import LOCAL_DATAMODEL_DIR, ensure_output_dir
from xml_utils import (
    strip_ns, get_description_text, build_type_string,
    get_syntax_type, get_default_value,
)

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
HEADER_FONT = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)

OBJECT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
PARAM_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
COMMAND_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
EVENT_FILL = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
ARG_CONTAINER_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
ARG_PARAM_FILL = PatternFill(start_color='FAFAFA', end_color='FAFAFA', fill_type='solid')
OBSOLETED_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
DEPRECATED_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


# ---------------------------------------------------------------------------
# XML processing
# ---------------------------------------------------------------------------

def process_xml(xml_path):
    """Parse the XML and extract all rows for the Device-2 sheet."""
    tree = ET.parse(xml_path)
    root = tree.getroot()

    model = None
    for child in root:
        if strip_ns(child.tag) == 'model':
            model = child
    if model is None:
        print("ERROR: No <model> element found in XML.", file=sys.stderr)
        sys.exit(1)

    model_name = model.attrib.get('name', 'Device:2')
    print(f"Processing model: {model_name}")

    rows = []
    for obj in model:
        tag = strip_ns(obj.tag)
        if tag != 'object':
            continue

        obj_name = obj.attrib.get('name', '')
        obj_access = obj.attrib.get('access', 'readOnly')
        obj_version = obj.attrib.get('version', '')
        obj_desc = get_description_text(obj.find('description'))
        obj_status = obj.attrib.get('status', '')

        write = 'W' if obj_access == 'readWrite' else 'R'
        rows.append({
            'row_type': 'object', 'name': obj_name,
            'type': 'object', 'write': write,
            'description': obj_desc, 'default': '-',
            'version': obj_version, 'status': obj_status,
        })

        for child in obj:
            ctag = strip_ns(child.tag)
            if ctag == 'parameter':
                _add_parameter_row(rows, child, 'parameter')
            elif ctag == 'command':
                _add_command_rows(rows, child)
            elif ctag == 'event':
                _add_event_rows(rows, child)

    print(f"Total rows: {len(rows)}")
    return model_name, rows


def _add_parameter_row(rows, param_elem, row_type, prefix=''):
    name = param_elem.attrib.get('name', '')
    access = param_elem.attrib.get('access', 'readOnly')
    version = param_elem.attrib.get('version', '')
    status = param_elem.attrib.get('status', '')
    desc = get_description_text(param_elem.find('description'))
    type_str = get_syntax_type(param_elem)
    default_val = get_default_value(param_elem)
    write = 'W' if access == 'readWrite' else 'R'
    display_name = f'{prefix}{name}' if prefix else name
    rows.append({
        'row_type': row_type, 'name': display_name,
        'type': type_str, 'write': write,
        'description': desc, 'default': default_val,
        'version': version, 'status': status,
    })


def _add_command_rows(rows, cmd_elem):
    name = cmd_elem.attrib.get('name', '')
    version = cmd_elem.attrib.get('version', '')
    status = cmd_elem.attrib.get('status', '')
    desc = get_description_text(cmd_elem.find('description'))
    rows.append({
        'row_type': 'command', 'name': name,
        'type': 'command', 'write': '-',
        'description': desc, 'default': '-',
        'version': version, 'status': status,
    })
    inp = cmd_elem.find('input')
    if inp is not None:
        _add_argument_section(rows, inp, 'Input')
    out = cmd_elem.find('output')
    if out is not None:
        _add_argument_section(rows, out, 'Output')


def _add_event_rows(rows, event_elem):
    name = event_elem.attrib.get('name', '')
    version = event_elem.attrib.get('version', '')
    status = event_elem.attrib.get('status', '')
    desc = get_description_text(event_elem.find('description'))
    rows.append({
        'row_type': 'event', 'name': name,
        'type': 'event', 'write': '-',
        'description': desc, 'default': '-',
        'version': version, 'status': status,
    })
    has_params = False
    for child in event_elem:
        ctag = strip_ns(child.tag)
        if ctag == 'parameter':
            if not has_params:
                rows.append({
                    'row_type': 'argument-container', 'name': '\u21d2 Output.',
                    'type': 'arguments', 'write': '-',
                    'description': 'Output arguments.', 'default': '-',
                    'version': '', 'status': '',
                })
                has_params = True
            _add_argument_param_row(rows, child, 'R')
        elif ctag == 'object':
            if not has_params:
                rows.append({
                    'row_type': 'argument-container', 'name': '\u21d2 Output.',
                    'type': 'arguments', 'write': '-',
                    'description': 'Output arguments.', 'default': '-',
                    'version': '', 'status': '',
                })
                has_params = True
            _add_argument_object(rows, child)


def _add_argument_section(rows, section_elem, section_name):
    rows.append({
        'row_type': 'argument-container', 'name': f'\u21d2 {section_name}.',
        'type': 'arguments', 'write': '-',
        'description': f'{section_name} arguments.', 'default': '-',
        'version': '', 'status': '',
    })
    default_write = 'W' if section_name == 'Input' else 'R'
    for child in section_elem:
        ctag = strip_ns(child.tag)
        if ctag == 'parameter':
            _add_argument_param_row(rows, child, default_write)
        elif ctag == 'object':
            _add_argument_object(rows, child)


def _add_argument_param_row(rows, param_elem, default_write):
    name = param_elem.attrib.get('name', '')
    version = param_elem.attrib.get('version', '')
    status = param_elem.attrib.get('status', '')
    desc = get_description_text(param_elem.find('description'))
    type_str = get_syntax_type(param_elem)
    default_val = get_default_value(param_elem)
    access = param_elem.attrib.get('access', '')
    write = ('W' if access == 'readWrite' else 'R') if access else default_write
    rows.append({
        'row_type': 'argument-parameter', 'name': f'\u21d2 {name}',
        'type': type_str, 'write': write,
        'description': desc, 'default': default_val,
        'version': version, 'status': status,
    })


def _add_argument_object(rows, obj_elem):
    name = obj_elem.attrib.get('name', '')
    version = obj_elem.attrib.get('version', '')
    status = obj_elem.attrib.get('status', '')
    desc = get_description_text(obj_elem.find('description'))
    access = obj_elem.attrib.get('access', 'readOnly')
    write = 'W' if access == 'readWrite' else 'R'
    rows.append({
        'row_type': 'argument-object', 'name': f'\u21d2 {name}',
        'type': 'object', 'write': write,
        'description': desc, 'default': '-',
        'version': version, 'status': status,
    })
    for child in obj_elem:
        if strip_ns(child.tag) == 'parameter':
            pname = child.attrib.get('name', '')
            pversion = child.attrib.get('version', '')
            pstatus = child.attrib.get('status', '')
            pdesc = get_description_text(child.find('description'))
            ptype = get_syntax_type(child)
            pdefault = get_default_value(child)
            paccess = child.attrib.get('access', 'readOnly')
            pwrite = 'W' if paccess == 'readWrite' else 'R'
            rows.append({
                'row_type': 'argument-parameter', 'name': f'\u21d2 {pname}',
                'type': ptype, 'write': pwrite,
                'description': pdesc, 'default': pdefault,
                'version': pversion, 'status': pstatus,
            })


# ---------------------------------------------------------------------------
# Data Types sheet
# ---------------------------------------------------------------------------

def extract_data_types(xml_path):
    """Extract all <dataType> definitions from the XML."""
    tree = ET.parse(xml_path)
    root = tree.getroot()
    datatypes = []
    for elem in root.iter():
        if strip_ns(elem.tag) == 'dataType':
            name = elem.attrib.get('name', '')
            if not name:
                continue
            base = elem.attrib.get('base', '')
            desc = get_description_text(elem.find('description'))
            datatypes.append({'name': name, 'base': base, 'description': desc})
    return datatypes


# ---------------------------------------------------------------------------
# References sheet
# ---------------------------------------------------------------------------

def extract_references(xml_path):
    """Extract all <reference> entries from <bibliography>."""
    tree = ET.parse(xml_path)
    root = tree.getroot()
    refs = []
    for elem in root.iter():
        if strip_ns(elem.tag) == 'reference':
            ref_id = elem.attrib.get('id', '')
            name = ''
            title = ''
            hyperlink = ''
            for child in elem:
                ctag = strip_ns(child.tag)
                if ctag == 'name':
                    name = (child.text or '').strip()
                elif ctag == 'title':
                    title = (child.text or '').strip()
                elif ctag == 'hyperlink':
                    hyperlink = (child.text or '').strip()
            refs.append({'id': ref_id, 'name': name, 'title': title, 'url': hyperlink})
    return refs


# ---------------------------------------------------------------------------
# Excel generation
# ---------------------------------------------------------------------------

def _object_path(name, row_type):
    """Extract the object path (parent) from a row name."""
    if row_type == 'object':
        return name
    return ''


def generate_excel(model_name, rows, output_path, xml_path):
    """Generate the complete Excel workbook."""
    wb = Workbook()

    # --- Device-2 sheet ---
    ws = wb.active
    ws.title = "Device-2"

    col_widths = {
        'A': 40, 'B': 40, 'C': 20, 'D': 18, 'E': 7,
        'F': 80, 'G': 16, 'H': 9,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    headers = ['Object Path', 'Item Path', 'Name', 'Type', 'Write',
               'Description', 'Object Default', 'Version']
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:H{len(rows) + 1}'

    current_object_path = ''
    current_cmd_event = ''  # tracks the current command/event for argument paths
    for row_idx, row_data in enumerate(rows, 2):
        row_type = row_data['row_type']
        status = row_data.get('status', '')
        name = row_data['name']

        # Compute Object Path and Item Path
        if row_type == 'object':
            current_object_path = name
            current_cmd_event = ''
            obj_path = name
            item_path = name
        elif row_type in ('command', 'event'):
            obj_path = current_object_path
            item_path = current_object_path + name
            current_cmd_event = item_path
        elif row_type == 'parameter':
            obj_path = current_object_path
            item_path = current_object_path + name
        elif row_type == 'argument-container':
            # e.g. "⇒ Input." → "Device.X.Backup()⇒Input."
            obj_path = current_object_path
            arg_label = name.replace('\u21d2 ', '')  # "Input." or "Output."
            item_path = current_cmd_event + '\u21d2' + arg_label if current_cmd_event else ''
        elif row_type in ('argument-parameter', 'argument-object'):
            # e.g. "⇒ URL" → "Device.X.Backup()⇒URL"
            obj_path = current_object_path
            arg_label = name.replace('\u21d2 ', '')
            item_path = current_cmd_event + '\u21d2' + arg_label if current_cmd_event else ''
        else:
            obj_path = current_object_path
            item_path = ''

        # Determine fill color
        if status == 'obsoleted':
            fill = OBSOLETED_FILL
        elif status == 'deprecated':
            fill = DEPRECATED_FILL
        elif row_type == 'object':
            fill = OBJECT_FILL
        elif row_type == 'command':
            fill = COMMAND_FILL
        elif row_type == 'event':
            fill = EVENT_FILL
        elif row_type == 'argument-container':
            fill = ARG_CONTAINER_FILL
        elif row_type in ('argument-parameter', 'argument-object'):
            fill = ARG_PARAM_FILL
        else:
            fill = PARAM_FILL

        # A: Object Path
        ws.cell(row=row_idx, column=1, value=obj_path)
        # B: Item Path
        ws.cell(row=row_idx, column=2, value=item_path)
        # C: Name
        name_cell = ws.cell(row=row_idx, column=3, value=name)
        if row_type in ('object', 'argument-object'):
            name_cell.font = Font(name='Calibri', bold=True, size=10)
        elif row_type == 'command':
            name_cell.font = Font(name='Calibri', bold=True, size=10, color='2E7D32')
        elif row_type == 'event':
            name_cell.font = Font(name='Calibri', bold=True, size=10, color='E65100')
        elif row_type == 'argument-container':
            name_cell.font = Font(name='Calibri', italic=True, size=10, color='666666')
        elif row_type == 'argument-parameter':
            name_cell.font = Font(name='Calibri', size=10, color='444444')
        elif status == 'obsoleted':
            name_cell.font = Font(name='Calibri', size=10, color='999999', strikethrough=True)
        elif status == 'deprecated':
            name_cell.font = Font(name='Calibri', size=10, color='B8860B')
        else:
            name_cell.font = Font(name='Calibri', size=10)

        # D: Type
        ws.cell(row=row_idx, column=4, value=row_data['type']).font = Font(
            name='Calibri', size=10, color='555555')
        # E: Write
        write_cell = ws.cell(row=row_idx, column=5, value=row_data['write'])
        write_cell.font = Font(name='Calibri', size=10)
        write_cell.alignment = Alignment(horizontal='center', vertical='top')
        # F: Description
        desc_text = row_data['description']
        if len(desc_text) > 32000:
            desc_text = desc_text[:32000] + '... [truncated]'
        desc_cell = ws.cell(row=row_idx, column=6, value=desc_text)
        desc_cell.font = Font(name='Calibri', size=9)
        desc_cell.alignment = Alignment(wrap_text=True, vertical='top')
        # G: Default
        default_cell = ws.cell(row=row_idx, column=7, value=row_data['default'])
        default_cell.font = Font(name='Calibri', size=10)
        default_cell.alignment = Alignment(horizontal='center', vertical='top')
        # H: Version
        version_cell = ws.cell(row=row_idx, column=8, value=row_data['version'])
        version_cell.font = Font(name='Calibri', size=10)
        version_cell.alignment = Alignment(horizontal='center', vertical='top')

        # Apply fill and border
        for col_idx in range(1, 9):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            if col_idx not in (6,):
                cell.alignment = Alignment(vertical='top', wrap_text=(col_idx in (1, 2, 3)))

        if row_idx % 5000 == 0:
            print(f"  Written {row_idx - 1} rows...")

    # --- Data Types sheet ---
    datatypes = extract_data_types(xml_path)
    if datatypes:
        ws_dt = wb.create_sheet('Data Types')
        dt_headers = ['Name', 'Base Type', 'Description']
        for ci, h in enumerate(dt_headers, 1):
            cell = ws_dt.cell(row=1, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        for ri, dt in enumerate(datatypes, 2):
            ws_dt.cell(row=ri, column=1, value=dt['name']).font = Font(name='Calibri', size=10)
            ws_dt.cell(row=ri, column=2, value=dt['base']).font = Font(name='Calibri', size=10)
            desc = dt['description']
            if len(desc) > 32000:
                desc = desc[:32000] + '... [truncated]'
            ws_dt.cell(row=ri, column=3, value=desc).font = Font(name='Calibri', size=9)
            for ci in range(1, 4):
                ws_dt.cell(row=ri, column=ci).border = THIN_BORDER
        ws_dt.column_dimensions['A'].width = 30
        ws_dt.column_dimensions['B'].width = 20
        ws_dt.column_dimensions['C'].width = 90
        ws_dt.freeze_panes = 'A2'

    # --- References sheet ---
    refs = extract_references(xml_path)
    if refs:
        ws_ref = wb.create_sheet('References')
        ref_headers = ['ID', 'Name', 'Title', 'URL']
        for ci, h in enumerate(ref_headers, 1):
            cell = ws_ref.cell(row=1, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = HEADER_ALIGN
            cell.border = THIN_BORDER
        for ri, ref in enumerate(refs, 2):
            ws_ref.cell(row=ri, column=1, value=ref['id']).font = Font(name='Calibri', size=10)
            ws_ref.cell(row=ri, column=2, value=ref['name']).font = Font(name='Calibri', size=10)
            ws_ref.cell(row=ri, column=3, value=ref['title']).font = Font(name='Calibri', size=10)
            ws_ref.cell(row=ri, column=4, value=ref['url']).font = Font(name='Calibri', size=10, color='0066CC')
            for ci in range(1, 5):
                ws_ref.cell(row=ri, column=ci).border = THIN_BORDER
        ws_ref.column_dimensions['A'].width = 20
        ws_ref.column_dimensions['B'].width = 30
        ws_ref.column_dimensions['C'].width = 60
        ws_ref.column_dimensions['D'].width = 50
        ws_ref.freeze_panes = 'A2'

    # --- Legend sheet ---
    ws_legend = wb.create_sheet('Legend')
    legend_data = [
        ['Row Color', 'Meaning'],
        ['Light Blue', 'Object definition'],
        ['White', 'Parameter definition'],
        ['Light Green', 'Command definition'],
        ['Light Orange', 'Event definition'],
        ['Light Gray', 'Argument container (Input/Output)'],
        ['Very Light Gray', 'Argument parameter'],
        ['Gray', 'Obsoleted item'],
        ['Light Yellow', 'Deprecated item'],
        ['', ''],
        ['Column', 'Description'],
        ['Object Path', 'Full dotted path of the parent object'],
        ['Item Path', 'Full dotted path of this item (object, param, command, event)'],
        ['Name', 'Full path for objects, short name for parameters. Commands end with (), events with !'],
        ['Type', 'Data type. string(:N) = max length N. type[] = comma-separated list. object = container.'],
        ['Write', 'R = read-only, W = read-write, - = not applicable'],
        ['Description', 'Full description from the data model specification'],
        ['Object Default', 'Default value. - means no default specified. <Empty> means empty string default.'],
        ['Version', 'Data model version when this item was introduced'],
    ]
    legend_fills = [
        HEADER_FILL, OBJECT_FILL, PARAM_FILL, COMMAND_FILL, EVENT_FILL,
        ARG_CONTAINER_FILL, ARG_PARAM_FILL, OBSOLETED_FILL, DEPRECATED_FILL,
        None, HEADER_FILL,
        None, None, None, None, None, None, None, None,
    ]
    for r_idx, row_vals in enumerate(legend_data, 1):
        lfill = legend_fills[r_idx - 1] if r_idx - 1 < len(legend_fills) else None
        for c_idx, val in enumerate(row_vals, 1):
            cell = ws_legend.cell(row=r_idx, column=c_idx, value=val)
            if r_idx in (1, 11):
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
            elif lfill:
                cell.fill = lfill
    ws_legend.column_dimensions['A'].width = 20
    ws_legend.column_dimensions['B'].width = 90

    wb.save(output_path)
    print(f"Excel file saved: {output_path}")
    print(f"  Sheets: {', '.join(wb.sheetnames)}")
    print(f"  Device-2 rows: {len(rows)}")
    if datatypes:
        print(f"  Data Types: {len(datatypes)}")
    if refs:
        print(f"  References: {len(refs)}")


def main():
    parser = argparse.ArgumentParser(description='Generate Excel from TR-181 XML')
    parser.add_argument('--input', '-i', help='Input XML file path')
    parser.add_argument('--output-dir', '-o', default=LOCAL_DATAMODEL_DIR,
                        help='Output directory')
    args = parser.parse_args()

    ensure_output_dir(args.output_dir)

    # Auto-detect input file if not specified
    xml_path = args.input
    if not xml_path:
        # Look for full XML in local/datamodel/
        for fname in sorted(os.listdir(LOCAL_DATAMODEL_DIR), reverse=True):
            if fname.startswith('tr-181-') and fname.endswith('-usp-full.xml'):
                xml_path = os.path.join(LOCAL_DATAMODEL_DIR, fname)
                break
        if not xml_path:
            for fname in sorted(os.listdir(LOCAL_DATAMODEL_DIR), reverse=True):
                if fname.startswith('tr-181-') and fname.endswith('-usp.xml'):
                    xml_path = os.path.join(LOCAL_DATAMODEL_DIR, fname)
                    break
    if not xml_path or not os.path.exists(xml_path):
        print("ERROR: No input XML file found. Run step1 first or specify --input.", file=sys.stderr)
        sys.exit(1)

    output_path = os.path.join(args.output_dir, 'tr181_definition.xlsx')
    print(f"Input:  {xml_path}")
    print(f"Output: {output_path}")

    model_name, rows = process_xml(xml_path)
    generate_excel(model_name, rows, output_path, xml_path)


if __name__ == '__main__':
    main()
