#!/usr/bin/env python3
"""
Verification script for the GW3 data model report pipeline.

Validates pipeline output across 6 categories (~25 checks):
  A — File Integrity (INT-001 to INT-006)
  B — Statistics Consistency (STAT-001 to STAT-005)
  C — Spot Checks (SPOT-001 to SPOT-009)
  D — Proxy Mapping (PROXY-001 to PROXY-003)
  E — Cross-Format Consistency (XFMT-001 to XFMT-003)
  F — Parent-Child Consistency (TREE-001 to TREE-002)

Runs after steps 1-6 (before or after step 7).
Returns exit code 0 only if all critical checks pass.

Usage:
    python3 step_verify.py [--input-dir DIR] [--staging-dir DIR]
                           [--strict] [--json-output] [--verbose]
"""

import argparse
import json
import os
import sys
import xml.etree.ElementTree as ET
from dataclasses import dataclass, field, asdict
from typing import List, Optional

# Allow running from project root or from scripts/
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, SCRIPTS_DIR)

from config import (
    SPECS_DATAMODEL_DIR, LOCAL_DATAMODEL_DIR,
    GW3_NAMESPACE, staging_amx_path,
)
from xml_utils import strip_ns, normalize_path


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class CheckResult:
    check_id: str
    category: str
    status: str          # PASS, WARN, FAIL, SKIP
    message: str
    details: str = ""


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

GW3_NS = f'{{{GW3_NAMESPACE}}}'

REQUIRED_FILES = [
    'gw3_current_datamodel.xml',
    'gw3_current_datamodel.xlsx',
    'gw3_current_datamodel.json',
    'tr181_definition.xlsx',
    'tr181_original.json',
    'MANIFEST.md',
]

ENHANCED_EXCEL_HEADERS = [
    'Object Path', 'Item Path', 'Name', 'Type', 'Write',
    'Description', 'Object Default', 'Version',
    'Microservice', 'Mark', 'ODL Type', 'ODL Access', 'ODL Persistent',
    'ODL Volatile', 'ODL Default', 'ODL Source File', 'Microservice Version',
    'Diff Details',
]

VALID_MARKS = {
    'IMPLEMENTED', 'NOT IMPLEMENTED', 'DEFAULT ONLY', 'REPORT ONLY',
    'NOT DEFINED', 'MISMATCH', 'VENDOR EXTENSION', 'PRPL EXTENSION',
}


# ---------------------------------------------------------------------------
# Category A: File Integrity
# ---------------------------------------------------------------------------

def check_file_integrity(input_dir: str) -> List[CheckResult]:
    results = []

    # INT-001: All required files exist
    missing = [f for f in REQUIRED_FILES if not os.path.exists(os.path.join(input_dir, f))]
    if missing:
        results.append(CheckResult(
            'INT-001', 'File Integrity', 'FAIL',
            f'{len(missing)} required file(s) missing',
            ', '.join(missing),
        ))
    else:
        results.append(CheckResult(
            'INT-001', 'File Integrity', 'PASS',
            f'All {len(REQUIRED_FILES)} required files present',
        ))

    # INT-002: No zero-byte files
    zero_files = []
    for f in REQUIRED_FILES:
        fpath = os.path.join(input_dir, f)
        if os.path.exists(fpath) and os.path.getsize(fpath) == 0:
            zero_files.append(f)
    if zero_files:
        results.append(CheckResult(
            'INT-002', 'File Integrity', 'FAIL',
            f'{len(zero_files)} zero-byte file(s)',
            ', '.join(zero_files),
        ))
    else:
        results.append(CheckResult(
            'INT-002', 'File Integrity', 'PASS',
            'No zero-byte files',
        ))

    # INT-003: XML files parse
    xml_file = os.path.join(input_dir, 'gw3_current_datamodel.xml')
    if os.path.exists(xml_file):
        try:
            ET.parse(xml_file)
            results.append(CheckResult(
                'INT-003', 'File Integrity', 'PASS',
                'Enhanced XML parses successfully',
            ))
        except ET.ParseError as e:
            results.append(CheckResult(
                'INT-003', 'File Integrity', 'FAIL',
                'Enhanced XML parse error',
                str(e),
            ))
    else:
        results.append(CheckResult(
            'INT-003', 'File Integrity', 'SKIP',
            'Enhanced XML not found',
        ))

    # INT-004: JSON files parse
    json_files = ['gw3_current_datamodel.json', 'tr181_original.json']
    json_ok = True
    json_details = []
    for jf in json_files:
        jpath = os.path.join(input_dir, jf)
        if os.path.exists(jpath):
            try:
                with open(jpath) as f:
                    json.load(f)
            except (json.JSONDecodeError, ValueError) as e:
                json_ok = False
                json_details.append(f'{jf}: {e}')
    if json_ok:
        results.append(CheckResult(
            'INT-004', 'File Integrity', 'PASS',
            'All JSON files parse successfully',
        ))
    else:
        results.append(CheckResult(
            'INT-004', 'File Integrity', 'FAIL',
            'JSON parse error(s)',
            '; '.join(json_details),
        ))

    # INT-005: Enhanced XML has gw3 namespace
    if os.path.exists(xml_file):
        try:
            tree = ET.parse(xml_file)
            root = tree.getroot()
            # Check if any element has gw3 namespace attributes
            has_gw3 = False
            model = None
            for child in root:
                if strip_ns(child.tag) == 'model':
                    model = child
                    break
            if model is not None:
                for obj in model:
                    if strip_ns(obj.tag) != 'object':
                        continue
                    for attr_key in obj.attrib:
                        if GW3_NAMESPACE in attr_key:
                            has_gw3 = True
                            break
                    if has_gw3:
                        break
            if has_gw3:
                results.append(CheckResult(
                    'INT-005', 'File Integrity', 'PASS',
                    f'Enhanced XML has gw3 namespace ({GW3_NAMESPACE})',
                ))
            else:
                results.append(CheckResult(
                    'INT-005', 'File Integrity', 'FAIL',
                    'Enhanced XML missing gw3 namespace attributes',
                ))
        except ET.ParseError:
            results.append(CheckResult(
                'INT-005', 'File Integrity', 'SKIP',
                'Cannot check namespace — XML parse failed',
            ))
    else:
        results.append(CheckResult(
            'INT-005', 'File Integrity', 'SKIP',
            'Enhanced XML not found',
        ))

    # INT-006: Enhanced Excel has 18 columns
    xlsx_path = os.path.join(input_dir, 'gw3_current_datamodel.xlsx')
    if os.path.exists(xlsx_path):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)
            # Find Device-2 sheet
            if 'Device-2' in wb.sheetnames:
                ws = wb['Device-2']
                headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                if len(headers) >= 18:
                    # Check header names match
                    mismatched = []
                    for i, (got, expected) in enumerate(zip(headers, ENHANCED_EXCEL_HEADERS)):
                        if got != expected:
                            mismatched.append(f'col{i+1}: got "{got}", expected "{expected}"')
                    if not mismatched:
                        results.append(CheckResult(
                            'INT-006', 'File Integrity', 'PASS',
                            '18 columns with correct headers',
                        ))
                    else:
                        results.append(CheckResult(
                            'INT-006', 'File Integrity', 'FAIL',
                            f'{len(mismatched)} header mismatch(es)',
                            '; '.join(mismatched[:5]),
                        ))
                else:
                    results.append(CheckResult(
                        'INT-006', 'File Integrity', 'FAIL',
                        f'Only {len(headers)} columns (expected 18)',
                    ))
            else:
                results.append(CheckResult(
                    'INT-006', 'File Integrity', 'FAIL',
                    'Device-2 sheet not found in enhanced Excel',
                ))
            wb.close()
        except ImportError:
            results.append(CheckResult(
                'INT-006', 'File Integrity', 'SKIP',
                'openpyxl not installed — cannot verify Excel',
            ))
        except Exception as e:
            results.append(CheckResult(
                'INT-006', 'File Integrity', 'FAIL',
                f'Error reading Excel: {e}',
            ))
    else:
        results.append(CheckResult(
            'INT-006', 'File Integrity', 'SKIP',
            'Enhanced Excel not found',
        ))

    return results


# ---------------------------------------------------------------------------
# Helper: Parse enhanced XML into mark/path data
# ---------------------------------------------------------------------------

def _parse_enhanced_xml(input_dir: str):
    """Parse the enhanced XML and return (elements, stats) or (None, None)."""
    xml_path = os.path.join(input_dir, 'gw3_current_datamodel.xml')
    if not os.path.exists(xml_path):
        return None, None

    tree = ET.parse(xml_path)
    root = tree.getroot()

    model = None
    for child in root:
        if strip_ns(child.tag) == 'model':
            model = child
            break
    if model is None:
        return None, None

    elements = []  # list of (path, tag, mark, service, attrs_dict)
    stats = {'total': 0}

    for obj in model:
        if strip_ns(obj.tag) != 'object':
            continue

        obj_name = obj.attrib.get('name', '')
        mark = obj.attrib.get(f'{GW3_NS}Mark', '')
        svc = obj.attrib.get(f'{GW3_NS}Microservice', '')

        elements.append((obj_name, 'object', mark, svc, dict(obj.attrib)))
        if mark:
            stats[mark] = stats.get(mark, 0) + 1
            stats['total'] += 1

        for child in obj:
            ctag = strip_ns(child.tag)
            if ctag not in ('parameter', 'command', 'event'):
                continue
            c_name = child.attrib.get('name', '')
            c_mark = child.attrib.get(f'{GW3_NS}Mark', '')
            c_svc = child.attrib.get(f'{GW3_NS}Microservice', '')

            if ctag == 'parameter':
                full_path = obj_name + c_name
            elif ctag == 'command':
                full_path = obj_name + c_name
            else:
                full_path = obj_name + c_name

            elements.append((full_path, ctag, c_mark, c_svc, dict(child.attrib)))
            if c_mark:
                stats[c_mark] = stats.get(c_mark, 0) + 1
                stats['total'] += 1

    return elements, stats


# ---------------------------------------------------------------------------
# Category B: Statistics Consistency
# ---------------------------------------------------------------------------

def check_statistics(input_dir: str) -> List[CheckResult]:
    results = []
    elements, stats = _parse_enhanced_xml(input_dir)

    if elements is None:
        results.append(CheckResult(
            'STAT-001', 'Statistics', 'SKIP', 'Enhanced XML not available'))
        return results

    total = stats.get('total', 0)

    # STAT-001: Total annotated items >= 5,000
    if total < 5000:
        results.append(CheckResult(
            'STAT-001', 'Statistics', 'FAIL',
            f'Total annotated items: {total} (need >= 5,000)',
        ))
    elif total < 8000:
        results.append(CheckResult(
            'STAT-001', 'Statistics', 'WARN',
            f'Total annotated items: {total} (expected >= 8,000)',
        ))
    else:
        results.append(CheckResult(
            'STAT-001', 'Statistics', 'PASS',
            f'Total annotated items: {total}',
        ))

    # STAT-002: Sum of all marks == total
    mark_sum = sum(v for k, v in stats.items() if k in VALID_MARKS)
    if mark_sum != total:
        results.append(CheckResult(
            'STAT-002', 'Statistics', 'FAIL',
            f'Mark sum ({mark_sum}) != total ({total})',
        ))
    else:
        results.append(CheckResult(
            'STAT-002', 'Statistics', 'PASS',
            f'Mark sum matches total ({total})',
        ))

    # STAT-003: IMPLEMENTED% between 3% and 60%
    impl_count = stats.get('IMPLEMENTED', 0)
    impl_pct = impl_count / total * 100 if total else 0
    if impl_pct < 3 or impl_pct > 60:
        results.append(CheckResult(
            'STAT-003', 'Statistics', 'FAIL',
            f'IMPLEMENTED rate: {impl_pct:.1f}% (expected 3-60%)',
            f'{impl_count} of {total}',
        ))
    else:
        results.append(CheckResult(
            'STAT-003', 'Statistics', 'PASS',
            f'IMPLEMENTED rate: {impl_pct:.1f}% ({impl_count} items)',
        ))

    # STAT-004: NOT DEFINED is the largest category
    not_def = stats.get('NOT DEFINED', 0)
    largest_mark = max(
        (k for k in VALID_MARKS if k in stats),
        key=lambda k: stats.get(k, 0),
        default='',
    )
    if largest_mark != 'NOT DEFINED':
        results.append(CheckResult(
            'STAT-004', 'Statistics', 'WARN',
            f'Largest category is {largest_mark} ({stats.get(largest_mark, 0)}), '
            f'expected NOT DEFINED ({not_def})',
        ))
    else:
        results.append(CheckResult(
            'STAT-004', 'Statistics', 'PASS',
            f'NOT DEFINED is largest category ({not_def})',
        ))

    # STAT-005: VENDOR EXTENSION count > 0
    vendor = stats.get('VENDOR EXTENSION', 0)
    prpl_ext = stats.get('PRPL EXTENSION', 0)
    ext_total = vendor + prpl_ext
    if ext_total == 0:
        results.append(CheckResult(
            'STAT-005', 'Statistics', 'FAIL',
            'No VENDOR EXTENSION or PRPL EXTENSION entries found',
        ))
    else:
        results.append(CheckResult(
            'STAT-005', 'Statistics', 'PASS',
            f'Extension entries: VENDOR EXTENSION={vendor}, PRPL EXTENSION={prpl_ext}',
        ))

    return results


# ---------------------------------------------------------------------------
# Category C: Spot Checks
# ---------------------------------------------------------------------------

def check_spot_checks(input_dir: str) -> List[CheckResult]:
    results = []
    elements, stats = _parse_enhanced_xml(input_dir)

    if elements is None:
        results.append(CheckResult(
            'SPOT-001', 'Spot Checks', 'SKIP', 'Enhanced XML not available'))
        return results

    # Build lookup by normalized path
    by_path = {}
    for path, tag, mark, svc, attrs in elements:
        norm = normalize_path(path)
        by_path[norm] = (path, tag, mark, svc, attrs)

    # SPOT-001: Device.IP.Interface. → IMPLEMENTED, ip-manager
    key = 'Device.IP.Interface.'
    if key in by_path:
        _, _, mark, svc, _ = by_path[key]
        if mark == 'IMPLEMENTED' and 'ip-manager' in svc:
            results.append(CheckResult(
                'SPOT-001', 'Spot Checks', 'PASS',
                f'{key} mark=IMPLEMENTED svc={svc}',
            ))
        else:
            results.append(CheckResult(
                'SPOT-001', 'Spot Checks', 'FAIL',
                f'{key} mark={mark} svc={svc} (expected IMPLEMENTED/ip-manager)',
            ))
    else:
        results.append(CheckResult(
            'SPOT-001', 'Spot Checks', 'FAIL',
            f'{key} not found in XML',
        ))

    # SPOT-002: Device.DeviceInfo. → IMPLEMENTED
    key = 'Device.DeviceInfo.'
    if key in by_path:
        _, _, mark, _, _ = by_path[key]
        if mark == 'IMPLEMENTED':
            results.append(CheckResult(
                'SPOT-002', 'Spot Checks', 'PASS',
                f'{key} mark=IMPLEMENTED',
            ))
        else:
            results.append(CheckResult(
                'SPOT-002', 'Spot Checks', 'FAIL',
                f'{key} mark={mark} (expected IMPLEMENTED)',
            ))
    else:
        results.append(CheckResult(
            'SPOT-002', 'Spot Checks', 'FAIL',
            f'{key} not found',
        ))

    # SPOT-003: All Device.DSL.* → NOT DEFINED
    dsl_impl = [(p, m) for p, _, m, _, _ in elements if p.startswith('Device.DSL.') and m == 'IMPLEMENTED']
    if dsl_impl:
        results.append(CheckResult(
            'SPOT-003', 'Spot Checks', 'FAIL',
            f'{len(dsl_impl)} Device.DSL.* items marked IMPLEMENTED',
            ', '.join(p for p, _ in dsl_impl[:5]),
        ))
    else:
        results.append(CheckResult(
            'SPOT-003', 'Spot Checks', 'PASS',
            'No Device.DSL.* items marked IMPLEMENTED',
        ))

    # SPOT-004: All Device.MoCA.* → NOT DEFINED
    moca_impl = [(p, m) for p, _, m, _, _ in elements if p.startswith('Device.MoCA.') and m == 'IMPLEMENTED']
    if moca_impl:
        results.append(CheckResult(
            'SPOT-004', 'Spot Checks', 'FAIL',
            f'{len(moca_impl)} Device.MoCA.* items marked IMPLEMENTED',
            ', '.join(p for p, _ in moca_impl[:5]),
        ))
    else:
        results.append(CheckResult(
            'SPOT-004', 'Spot Checks', 'PASS',
            'No Device.MoCA.* items marked IMPLEMENTED',
        ))

    # SPOT-005: Device.Hosts. → IMPLEMENTED, hosts-manager
    key = 'Device.Hosts.'
    if key in by_path:
        _, _, mark, svc, _ = by_path[key]
        if mark == 'IMPLEMENTED':
            results.append(CheckResult(
                'SPOT-005', 'Spot Checks', 'PASS',
                f'{key} mark=IMPLEMENTED svc={svc}',
            ))
        else:
            results.append(CheckResult(
                'SPOT-005', 'Spot Checks', 'FAIL',
                f'{key} mark={mark} (expected IMPLEMENTED)',
            ))
    else:
        results.append(CheckResult(
            'SPOT-005', 'Spot Checks', 'FAIL',
            f'{key} not found',
        ))

    # SPOT-006: X_PRPLWARE-COM_* vendor extensions exist
    vendor_entries = [p for p, _, m, _, _ in elements if 'X_PRPLWARE-COM_' in p and m == 'VENDOR EXTENSION']
    if vendor_entries:
        results.append(CheckResult(
            'SPOT-006', 'Spot Checks', 'PASS',
            f'{len(vendor_entries)} X_PRPLWARE-COM_* vendor extensions found',
        ))
    else:
        results.append(CheckResult(
            'SPOT-006', 'Spot Checks', 'FAIL',
            'No X_PRPLWARE-COM_* vendor extensions found',
        ))

    # SPOT-007: Device.WiFi.* has entries from wld
    wifi_wld = [p for p, _, _, svc, _ in elements if p.startswith('Device.WiFi.') and svc and 'wld' in svc]
    if wifi_wld:
        results.append(CheckResult(
            'SPOT-007', 'Spot Checks', 'PASS',
            f'{len(wifi_wld)} Device.WiFi.* entries from wld',
        ))
    else:
        results.append(CheckResult(
            'SPOT-007', 'Spot Checks', 'WARN',
            'No Device.WiFi.* entries from wld',
        ))

    # SPOT-008: Device.SFPs.* entries exist (proxy-mapped)
    sfp_entries = [p for p, _, m, _, _ in elements if p.startswith('Device.SFPs.') and m != 'NOT DEFINED']
    if sfp_entries:
        results.append(CheckResult(
            'SPOT-008', 'Spot Checks', 'PASS',
            f'{len(sfp_entries)} Device.SFPs.* entries with non-NOT-DEFINED marks',
        ))
    else:
        # SFPs may not exist on all platforms
        sfp_any = [p for p, _, _, _, _ in elements if p.startswith('Device.SFPs.')]
        if sfp_any:
            results.append(CheckResult(
                'SPOT-008', 'Spot Checks', 'WARN',
                f'Device.SFPs.* exists ({len(sfp_any)} items) but all NOT DEFINED',
            ))
        else:
            results.append(CheckResult(
                'SPOT-008', 'Spot Checks', 'WARN',
                'No Device.SFPs.* entries found (may not exist in this TR-181 version)',
            ))

    # SPOT-009: Device.DeviceInfo.PowerStatus.* entries exist (proxy-mapped)
    ps_entries = [p for p, _, m, _, _ in elements if p.startswith('Device.DeviceInfo.PowerStatus.')]
    if ps_entries:
        non_nd = [p for p, _, m, _, _ in elements
                  if p.startswith('Device.DeviceInfo.PowerStatus.') and m != 'NOT DEFINED']
        if non_nd:
            results.append(CheckResult(
                'SPOT-009', 'Spot Checks', 'PASS',
                f'{len(non_nd)} Device.DeviceInfo.PowerStatus.* with non-NOT-DEFINED marks',
            ))
        else:
            results.append(CheckResult(
                'SPOT-009', 'Spot Checks', 'FAIL',
                f'{len(ps_entries)} Device.DeviceInfo.PowerStatus.* all NOT DEFINED (proxy mapping issue)',
            ))
    else:
        results.append(CheckResult(
            'SPOT-009', 'Spot Checks', 'FAIL',
            'No Device.DeviceInfo.PowerStatus.* entries found',
        ))

    return results


# ---------------------------------------------------------------------------
# Category D: Proxy Mapping
# ---------------------------------------------------------------------------

def check_proxy_mappings(input_dir: str, staging_dir: Optional[str]) -> List[CheckResult]:
    results = []

    if staging_dir is None or not os.path.isdir(staging_dir):
        results.append(CheckResult(
            'PROXY-001', 'Proxy Mapping', 'SKIP',
            'Staging directory not available — skipping proxy checks',
        ))
        return results

    try:
        from step3_generate_enhanced_xml import parse_proxy_mappings
        from service_map import build_service_map
    except ImportError:
        results.append(CheckResult(
            'PROXY-001', 'Proxy Mapping', 'SKIP',
            'Cannot import proxy/service_map modules',
        ))
        return results

    proxy_mappings = parse_proxy_mappings(staging_dir)

    # PROXY-001: Proxy mappings found > 0
    if not proxy_mappings:
        results.append(CheckResult(
            'PROXY-001', 'Proxy Mapping', 'FAIL',
            'No proxy mappings found in staging directory',
        ))
        return results
    else:
        results.append(CheckResult(
            'PROXY-001', 'Proxy Mapping', 'PASS',
            f'{len(proxy_mappings)} proxy mapping(s) found',
            ', '.join(f'{dp} -> {lp}' for dp, lp in sorted(proxy_mappings.items())),
        ))

    # Parse XML to check resolved entries
    elements, stats = _parse_enhanced_xml(input_dir)
    if elements is None:
        results.append(CheckResult(
            'PROXY-002', 'Proxy Mapping', 'SKIP',
            'Enhanced XML not available for cross-check',
        ))
        return results

    # PROXY-002: Each mapping resolves >= 1 entry in XML
    unresolved = []
    for device_prefix in sorted(proxy_mappings.keys()):
        matched = [p for p, _, m, _, _ in elements
                   if p.startswith(device_prefix) and m not in ('NOT DEFINED', '')]
        if not matched:
            unresolved.append(device_prefix)

    if unresolved:
        results.append(CheckResult(
            'PROXY-002', 'Proxy Mapping', 'WARN',
            f'{len(unresolved)} proxy mapping(s) with no resolved entries',
            ', '.join(unresolved),
        ))
    else:
        results.append(CheckResult(
            'PROXY-002', 'Proxy Mapping', 'PASS',
            'All proxy mappings resolve to >= 1 entry',
        ))

    # PROXY-003: Known proxy paths have non-NOT-DEFINED marks
    known_proxy_paths = [
        'Device.DeviceInfo.PowerStatus.',
    ]
    failed_known = []
    for kp in known_proxy_paths:
        entries = [p for p, _, m, _, _ in elements
                   if p.startswith(kp) and m not in ('NOT DEFINED', '')]
        if not entries:
            failed_known.append(kp)

    if failed_known:
        results.append(CheckResult(
            'PROXY-003', 'Proxy Mapping', 'FAIL',
            f'{len(failed_known)} known proxy path(s) have no non-NOT-DEFINED marks',
            ', '.join(failed_known),
        ))
    else:
        results.append(CheckResult(
            'PROXY-003', 'Proxy Mapping', 'PASS',
            'Known proxy paths have resolved entries',
        ))

    return results


# ---------------------------------------------------------------------------
# Category E: Cross-Format Consistency
# ---------------------------------------------------------------------------

def check_cross_format(input_dir: str) -> List[CheckResult]:
    results = []

    xml_path = os.path.join(input_dir, 'gw3_current_datamodel.xml')
    json_path = os.path.join(input_dir, 'gw3_current_datamodel.json')
    xlsx_path = os.path.join(input_dir, 'gw3_current_datamodel.xlsx')

    # Count objects in XML
    xml_obj_count = 0
    if os.path.exists(xml_path):
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            for child in root:
                if strip_ns(child.tag) == 'model':
                    for obj in child:
                        if strip_ns(obj.tag) == 'object':
                            xml_obj_count += 1
        except ET.ParseError:
            pass

    # XFMT-001: JSON object count within 5% of XML object count
    if os.path.exists(json_path) and xml_obj_count > 0:
        try:
            with open(json_path) as f:
                jdata = json.load(f)
            json_obj_count = len(jdata.get('objects', []))
            diff_pct = abs(json_obj_count - xml_obj_count) / xml_obj_count * 100
            if diff_pct > 5:
                results.append(CheckResult(
                    'XFMT-001', 'Cross-Format', 'FAIL',
                    f'JSON objects ({json_obj_count}) vs XML objects ({xml_obj_count}): {diff_pct:.1f}% diff',
                ))
            else:
                results.append(CheckResult(
                    'XFMT-001', 'Cross-Format', 'PASS',
                    f'JSON ({json_obj_count}) vs XML ({xml_obj_count}) objects: {diff_pct:.1f}% diff',
                ))
        except (json.JSONDecodeError, KeyError) as e:
            results.append(CheckResult(
                'XFMT-001', 'Cross-Format', 'SKIP',
                f'Error reading JSON: {e}',
            ))
    else:
        results.append(CheckResult(
            'XFMT-001', 'Cross-Format', 'SKIP',
            'JSON or XML not available for comparison',
        ))

    # XFMT-002: Excel data rows within 10% of XML annotated elements
    elements, stats = _parse_enhanced_xml(input_dir)
    xml_elem_count = stats.get('total', 0) if stats else 0

    if os.path.exists(xlsx_path) and xml_elem_count > 0:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True)
            if 'Device-2' in wb.sheetnames:
                ws = wb['Device-2']
                # Count data rows (excluding header)
                excel_rows = 0
                for row in ws.iter_rows(min_row=2, max_col=1):
                    if row[0].value is not None:
                        excel_rows += 1
                diff_pct = abs(excel_rows - xml_elem_count) / xml_elem_count * 100
                if diff_pct > 10:
                    results.append(CheckResult(
                        'XFMT-002', 'Cross-Format', 'FAIL',
                        f'Excel rows ({excel_rows}) vs XML elements ({xml_elem_count}): {diff_pct:.1f}% diff',
                    ))
                else:
                    results.append(CheckResult(
                        'XFMT-002', 'Cross-Format', 'PASS',
                        f'Excel rows ({excel_rows}) vs XML elements ({xml_elem_count}): {diff_pct:.1f}% diff',
                    ))
            else:
                results.append(CheckResult(
                    'XFMT-002', 'Cross-Format', 'SKIP',
                    'Device-2 sheet not found',
                ))
            wb.close()
        except ImportError:
            results.append(CheckResult(
                'XFMT-002', 'Cross-Format', 'SKIP',
                'openpyxl not installed',
            ))
    else:
        results.append(CheckResult(
            'XFMT-002', 'Cross-Format', 'SKIP',
            'Excel or XML not available for comparison',
        ))

    # XFMT-003: If both local/ and specs/ exist, file sizes match
    other_dir = None
    if input_dir.rstrip('/').endswith('specs/datamodel'):
        other_dir = input_dir.replace('specs/datamodel', 'local/datamodel')
    elif input_dir.rstrip('/').endswith('local/datamodel'):
        other_dir = input_dir.replace('local/datamodel', 'specs/datamodel')

    if other_dir and os.path.isdir(other_dir):
        mismatches = []
        common_files = ['gw3_current_datamodel.xml', 'gw3_current_datamodel.json']
        for cf in common_files:
            p1 = os.path.join(input_dir, cf)
            p2 = os.path.join(other_dir, cf)
            if os.path.exists(p1) and os.path.exists(p2):
                s1 = os.path.getsize(p1)
                s2 = os.path.getsize(p2)
                if s1 != s2:
                    mismatches.append(f'{cf}: {s1} vs {s2} bytes')
        if mismatches:
            results.append(CheckResult(
                'XFMT-003', 'Cross-Format', 'WARN',
                f'{len(mismatches)} file size mismatch(es) between dirs',
                '; '.join(mismatches),
            ))
        else:
            results.append(CheckResult(
                'XFMT-003', 'Cross-Format', 'PASS',
                'File sizes match between local/ and specs/',
            ))
    else:
        results.append(CheckResult(
            'XFMT-003', 'Cross-Format', 'SKIP',
            'Only one directory available — cannot compare',
        ))

    return results


# ---------------------------------------------------------------------------
# Category F: Parent-Child Consistency
# ---------------------------------------------------------------------------

def check_parent_child(input_dir: str) -> List[CheckResult]:
    results = []
    elements, stats = _parse_enhanced_xml(input_dir)

    if elements is None:
        results.append(CheckResult(
            'TREE-001', 'Parent-Child', 'SKIP',
            'Enhanced XML not available',
        ))
        return results

    # Build object mark lookup
    obj_marks = {}
    for path, tag, mark, svc, _ in elements:
        if tag == 'object':
            obj_marks[normalize_path(path)] = mark

    # TREE-001: IMPLEMENTED params under NOT DEFINED parent
    orphans = []
    for path, tag, mark, svc, _ in elements:
        if tag == 'parameter' and mark == 'IMPLEMENTED':
            # Find parent object
            parts = path.rsplit('.', 1)
            if len(parts) == 2:
                parent = parts[0] + '.'
                parent_norm = normalize_path(parent)
                parent_mark = obj_marks.get(parent_norm, '')
                if parent_mark == 'NOT DEFINED':
                    orphans.append(path)

    if len(orphans) > 100:
        results.append(CheckResult(
            'TREE-001', 'Parent-Child', 'FAIL',
            f'{len(orphans)} IMPLEMENTED params under NOT DEFINED parent',
            ', '.join(orphans[:10]) + '...',
        ))
    elif len(orphans) > 10:
        results.append(CheckResult(
            'TREE-001', 'Parent-Child', 'WARN',
            f'{len(orphans)} IMPLEMENTED params under NOT DEFINED parent',
            ', '.join(orphans[:10]),
        ))
    else:
        results.append(CheckResult(
            'TREE-001', 'Parent-Child', 'PASS',
            f'{len(orphans)} IMPLEMENTED params under NOT DEFINED parent (acceptable)',
        ))

    # TREE-002: IMPLEMENTED objects with no IMPLEMENTED children
    impl_objects = [path for path, tag, mark, _, _ in elements if tag == 'object' and mark == 'IMPLEMENTED']
    childless = []
    for obj_path in impl_objects:
        norm_obj = normalize_path(obj_path)
        has_impl_child = False
        for path, tag, mark, _, _ in elements:
            if tag != 'object' and mark == 'IMPLEMENTED':
                parent = path.rsplit('.', 1)[0] + '.' if '.' in path else ''
                if normalize_path(parent) == norm_obj:
                    has_impl_child = True
                    break
        if not has_impl_child:
            childless.append(obj_path)

    if childless:
        results.append(CheckResult(
            'TREE-002', 'Parent-Child', 'WARN',
            f'{len(childless)} IMPLEMENTED objects with no IMPLEMENTED children',
            ', '.join(childless[:10]),
        ))
    else:
        results.append(CheckResult(
            'TREE-002', 'Parent-Child', 'PASS',
            'All IMPLEMENTED objects have >= 1 IMPLEMENTED child',
        ))

    return results


# ---------------------------------------------------------------------------
# Output formatting
# ---------------------------------------------------------------------------

CATEGORY_NAMES = {
    'File Integrity': 'File Integrity',
    'Statistics': 'Statistics Consistency',
    'Spot Checks': 'Spot Checks',
    'Proxy Mapping': 'Proxy Mapping',
    'Cross-Format': 'Cross-Format Consistency',
    'Parent-Child': 'Parent-Child Consistency',
}


def print_results(all_results: List[CheckResult], verbose: bool = False):
    """Print results in human-readable format."""
    print("\n=== GW3 Data Model Report Verification ===\n")

    # Group by category
    by_cat = {}
    for r in all_results:
        by_cat.setdefault(r.category, []).append(r)

    for cat in ['File Integrity', 'Statistics', 'Spot Checks',
                'Proxy Mapping', 'Cross-Format', 'Parent-Child']:
        checks = by_cat.get(cat, [])
        if not checks:
            continue
        display = CATEGORY_NAMES.get(cat, cat)
        print(f"--- {display} ---")
        for r in checks:
            status_tag = f'[{r.status}]'
            print(f"  {status_tag:6s} {r.check_id}: {r.message}")
            if verbose and r.details:
                print(f"         {r.details}")
        print()


def print_summary(all_results: List[CheckResult], strict: bool) -> int:
    """Print summary and return exit code."""
    counts = {'PASS': 0, 'WARN': 0, 'FAIL': 0, 'SKIP': 0}
    for r in all_results:
        counts[r.status] = counts.get(r.status, 0) + 1

    effective_fail = counts['FAIL']
    if strict:
        effective_fail += counts['WARN']

    result = 'PASS' if effective_fail == 0 else 'FAIL'

    print("=== Summary ===")
    print(f"  PASS: {counts['PASS']}   WARN: {counts['WARN']}   "
          f"FAIL: {counts['FAIL']}   SKIP: {counts['SKIP']}")
    if strict and counts['WARN'] > 0:
        print(f"  (strict mode: {counts['WARN']} WARN treated as FAIL)")
    print(f"  Result: {result}")

    return 0 if effective_fail == 0 else 1


def print_json_results(all_results: List[CheckResult], strict: bool):
    """Print results as JSON."""
    output = {
        'checks': [asdict(r) for r in all_results],
        'summary': {
            'pass': sum(1 for r in all_results if r.status == 'PASS'),
            'warn': sum(1 for r in all_results if r.status == 'WARN'),
            'fail': sum(1 for r in all_results if r.status == 'FAIL'),
            'skip': sum(1 for r in all_results if r.status == 'SKIP'),
            'strict': strict,
        },
    }
    fail_count = output['summary']['fail']
    if strict:
        fail_count += output['summary']['warn']
    output['summary']['result'] = 'PASS' if fail_count == 0 else 'FAIL'
    print(json.dumps(output, indent=2))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description='Verify GW3 data model report pipeline output',
    )
    parser.add_argument(
        '--input-dir', '-d',
        help=f'Directory to verify (default: specs/datamodel/, falls back to local/datamodel/)',
    )
    parser.add_argument(
        '--staging-dir', '-s',
        help='AMX staging dir for proxy checks (default: auto-detect)',
    )
    parser.add_argument(
        '--strict', action='store_true',
        help='Treat WARNs as FAILs',
    )
    parser.add_argument(
        '--json-output', action='store_true',
        help='Output results as JSON',
    )
    parser.add_argument(
        '--verbose', '-v', action='store_true',
        help='Show detailed info for each check',
    )
    args = parser.parse_args()

    # Determine input directory
    input_dir = args.input_dir
    if not input_dir:
        if os.path.isdir(SPECS_DATAMODEL_DIR):
            input_dir = SPECS_DATAMODEL_DIR
        elif os.path.isdir(LOCAL_DATAMODEL_DIR):
            input_dir = LOCAL_DATAMODEL_DIR
        else:
            print("ERROR: No datamodel directory found. Run the pipeline first.",
                  file=sys.stderr)
            sys.exit(1)

    if not os.path.isdir(input_dir):
        print(f"ERROR: Directory not found: {input_dir}", file=sys.stderr)
        sys.exit(1)

    # Determine staging directory
    staging_dir = args.staging_dir
    if not staging_dir:
        default_staging = staging_amx_path()
        if os.path.isdir(default_staging):
            staging_dir = default_staging

    if not args.json_output:
        print(f"Input: {input_dir}")
        if staging_dir:
            print(f"Staging: {staging_dir}")
        else:
            print("Staging: not available (proxy checks will be skipped)")
        print()

    # Run all check categories
    all_results: List[CheckResult] = []
    all_results.extend(check_file_integrity(input_dir))
    all_results.extend(check_statistics(input_dir))
    all_results.extend(check_spot_checks(input_dir))
    all_results.extend(check_proxy_mappings(input_dir, staging_dir))
    all_results.extend(check_cross_format(input_dir))
    all_results.extend(check_parent_child(input_dir))

    # Output
    if args.json_output:
        print_json_results(all_results, args.strict)
        # Compute exit code without printing summary
        fail_count = sum(1 for r in all_results if r.status == 'FAIL')
        if args.strict:
            fail_count += sum(1 for r in all_results if r.status == 'WARN')
        sys.exit(0 if fail_count == 0 else 1)
    else:
        print_results(all_results, args.verbose)
        exit_code = print_summary(all_results, args.strict)
        sys.exit(exit_code)


if __name__ == '__main__':
    main()
